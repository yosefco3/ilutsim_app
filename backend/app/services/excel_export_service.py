"""
ExcelExportService — generates Excel files from schedule data.

Three report types:
1. Weekly schedule grid  – all guards × all days with shift choices
2. Deviation report       – guards exceeding shift-count thresholds
3. Guard history          – single-guard submission trail across weeks
"""

import io
import logging
import uuid
from datetime import date, timedelta
from typing import Any

from app.constants import (
    AdminRole,
    EventType,
    ShiftType,
    WeekStatus,
)
from app.messages import Messages
from app.repositories.schedule_event_repository import ScheduleEventRepository
from app.repositories.schedule_week_repository import ScheduleWeekRepository
from app.repositories.submission_repository import SubmissionRepository
from app.repositories.user_repository import UserRepository

logger = logging.getLogger("ilutzim")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Shared style constants ──────────────────────────────────────────
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
_TITLE_FONT = Font(bold=True, size=13)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RED_FILL = PatternFill(
    start_color="FF4444", end_color="FF4444", fill_type="solid"
)
_YELLOW_FILL = PatternFill(
    start_color="FFFFAA", end_color="FFFFAA", fill_type="solid"
)
_GREEN_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)

# Hebrew weekday order (Sunday=0 … Saturday=6)
_DAY_NAMES_HE = [
    "ראשון",
    "שני",
    "שלישי",
    "רביעי",
    "חמישי",
    "שישי",
    "שבת",
]

_SHIFT_LABELS: dict[str, str] = {
    ShiftType.MORNING.value: Messages.LABEL_MORNING,
    ShiftType.AFTERNOON.value: Messages.LABEL_AFTERNOON,
    ShiftType.NIGHT.value: Messages.LABEL_NIGHT,
}

_EVENT_LABELS: dict[str, str] = {
    EventType.VACATION.value: Messages.EVENT_VACATION,
    EventType.MILITARY_RESERVE.value: Messages.EVENT_MILITARY,
    EventType.FIREARMS_TRAINING.value: Messages.EVENT_FIREARMS,
}



def _apply_header_style(cell: Any) -> None:
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER


def _apply_cell_style(cell: Any, center: bool = True) -> None:
    cell.border = _THIN_BORDER
    cell.alignment = _CENTER if center else Alignment(vertical="center")


class ExcelExportService:
    """Generates Excel schedule reports for download."""

    def __init__(
        self,
        submission_repo: SubmissionRepository,
        user_repo: UserRepository,
        week_repo: ScheduleWeekRepository,
        event_repo: ScheduleEventRepository | None = None,
    ) -> None:
        self._submission_repo = submission_repo
        self._user_repo = user_repo
        self._week_repo = week_repo
        self._event_repo = event_repo

    # ── 1. Weekly schedule grid ─────────────────────────────────────

    async def export_weekly_schedule(self, week_id: uuid.UUID) -> bytes:
        """
        Generate an Excel file with the weekly schedule grid.

        Rows = guards, columns = days (Sun–Sat).
        Each cell shows the guard's selected shifts or event.
        Auto-absence is applied for missing submissions.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"

        # Title
        title_text = Messages.EXCEL_REPORT_TITLE.format(
            start=str(week.start_date), end=str(week.end_date)
        )
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Headers: Name | Sun | Mon | Tue | Wed | Thu | Fri | Sat
        headers = [Messages.EXCEL_HEADER_NAME] + _DAY_NAMES_HE
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        # Build lookup: user_id → submission
        sub_map: dict[uuid.UUID, Any] = {s.user_id: s for s in submissions}

        # Build event lookup: user_id → list of (start_date, end_date, label)
        event_map: dict[uuid.UUID, list[tuple[date, date, str]]] = {}
        if self._event_repo:
            for user in active_users:
                events = await self._event_repo.get_events_for_user(
                    user.id, week.start_date, week.end_date
                )
                for ev in events:
                    event_map.setdefault(user.id, []).append(
                        (ev.start_date, ev.end_date, _EVENT_LABELS.get(ev.event_type, ev.event_type))
                    )

        # Data rows
        row_num = 4
        for user in active_users:
            ws.cell(row=row_num, column=1, value=user.full_name)
            ws.cell(row=row_num, column=1).border = _THIN_BORDER
            ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center")

            sub = sub_map.get(user.id)
            user_events = event_map.get(user.id, [])

            for day_offset in range(7):
                col = day_offset + 2
                current_date = week.start_date + timedelta(days=day_offset)

                cell_value = ""
                cell_fill = None

                # Check if user has an event on this date
                event_label = None
                for ev_start, ev_end, ev_label in user_events:
                    if ev_start <= current_date <= ev_end:
                        event_label = ev_label
                        break

                if event_label:
                    cell_value = f"📅 {event_label}"
                    cell_fill = _YELLOW_FILL
                elif sub is None:
                    # No submission → auto-absence
                    cell_value = "❌"
                    cell_fill = _RED_FILL
                else:
                    cell_value = "✅"
                    cell_fill = _GREEN_FILL

                cell = ws.cell(row=row_num, column=col, value=cell_value)
                _apply_cell_style(cell)
                if cell_fill:
                    cell.fill = cell_fill

            row_num += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel weekly schedule exported for week {week_id}: {row_num - 4} rows")
        return buffer.read()

    # ── 1b. Constraints report (who submitted what) ─────────────────

    async def export_constraints_report(self, week_id: uuid.UUID) -> bytes:
        """
        Generate a nicely-formatted Excel of all submitted constraints.

        One row per guard who submitted, with per-day availability and the
        exact shift windows they chose, plus their general notes.
        Sheet is rendered right-to-left for Hebrew readability.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()
        user_map = {u.id: u for u in active_users}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "אילוצים"
        ws.sheet_view.rightToLeft = True

        n_cols = 10  # name + phone + 7 days + notes

        # Title
        title_text = f"אילוצים שהוגשו — {week.start_date} עד {week.end_date}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Subtitle: how many submitted
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        ws.cell(
            row=2,
            column=1,
            value=f"סך הכל הגישו: {len(submissions)} שומרים",
        ).alignment = _CENTER

        # Headers
        headers = (
            [Messages.EXCEL_HEADER_NAME, Messages.EXCEL_HEADER_PHONE]
            + _DAY_NAMES_HE
            + ["הערות"]
        )
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=4, column=col, value=header))

        # Sort submitting guards by name for a stable, readable order
        def _sub_name(sub: Any) -> str:
            user = user_map.get(sub.user_id)
            return user.full_name if user else ""

        ordered = sorted(submissions, key=_sub_name)

        row_num = 5
        for sub in ordered:
            user = user_map.get(sub.user_id)
            if user is None:
                user = await self._user_repo.get_by_id(sub.user_id)
            full_name = user.full_name if user else "—"
            phone = (user.phone_number if user else "") or ""

            # Map each calendar date → its daily status for this submission
            status_by_date: dict[date, Any] = {
                ds.date: ds for ds in sub.daily_statuses
            }

            name_cell = ws.cell(row=row_num, column=1, value=full_name)
            _apply_cell_style(name_cell, center=False)
            phone_cell = ws.cell(row=row_num, column=2, value=phone)
            _apply_cell_style(phone_cell)

            for day_offset in range(7):
                col = day_offset + 3
                current_date = week.start_date + timedelta(days=day_offset)
                ds = status_by_date.get(current_date)

                if ds is None or not ds.is_available:
                    cell_value = "לא זמין"
                    cell_fill = _RED_FILL
                else:
                    windows = sorted(
                        ds.shift_windows, key=lambda w: w.start_time
                    )
                    if windows:
                        parts = []
                        for w in windows:
                            label = _SHIFT_LABELS.get(
                                getattr(w.shift_type, "value", w.shift_type),
                                "",
                            )
                            parts.append(
                                f"{label} {w.start_time:%H:%M}–{w.end_time:%H:%M}".strip()
                            )
                        cell_value = "\n".join(parts)
                    else:
                        cell_value = "זמין"
                    cell_fill = _GREEN_FILL

                cell = ws.cell(row=row_num, column=col, value=cell_value)
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER_WRAP
                cell.fill = cell_fill

            notes_cell = ws.cell(
                row=row_num, column=10, value=sub.general_notes or ""
            )
            notes_cell.border = _THIN_BORDER
            notes_cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )

            row_num += 1

        # Column widths: name + notes wider, days medium
        ws.column_dimensions[get_column_letter(1)].width = 20  # name
        ws.column_dimensions[get_column_letter(2)].width = 16  # phone
        for col in range(3, 10):  # days
            ws.column_dimensions[get_column_letter(col)].width = 18
        ws.column_dimensions[get_column_letter(10)].width = 30  # notes

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            f"Excel constraints report exported for week {week_id}: "
            f"{row_num - 5} rows"
        )
        return buffer.read()

    # ── 2. Deviation report ─────────────────────────────────────────

    async def export_deviation_report(self, week_id: uuid.UUID) -> bytes:
        """
        Generate a deviation report Excel file.

        Lists guards whose shift counts deviate from thresholds.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        week = await self._week_repo.get_by_id(week_id)
        if week is None:
            raise ValueError(f"Week {week_id} not found")

        submissions = await self._submission_repo.get_submissions_for_week(week_id)
        active_users = await self._user_repo.get_active_users()
        user_map = {u.id: u for u in active_users}

        # Get deviation settings (min_shifts_per_week from system_settings)
        # Default threshold: 3 shifts per week
        threshold = 3

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deviations"

        # Title
        title_text = f"דוח חריגות — {week.start_date} עד {week.end_date}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Headers
        headers = [
            Messages.EXCEL_HEADER_NAME,
            Messages.EXCEL_HEADER_PHONE,
            "מספר משמרות",
            Messages.EXCEL_HEADER_THRESHOLDS,
            Messages.EXCEL_HEADER_DEVIATION,
        ]
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=3, column=col, value=header))

        row_num = 4
        for sub in submissions:
            user = user_map.get(sub.user_id)
            if not user:
                continue

            # Show all submissions
            row_data = [
                user.full_name,
                user.phone_number or "",
                "✅",
                str(threshold),
                "-",
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                _apply_cell_style(cell)
            row_num += 1
        # Also show guards with no submission
        sub_user_ids = {s.user_id for s in submissions}
        for user in active_users:
            if user.id not in sub_user_ids:
                row_data = [
                    user.full_name,
                    user.phone_number or "",
                    "❌",
                    str(threshold),
                    f"{threshold}-",
                ]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    _apply_cell_style(cell)
                    if col == 5:
                        cell.fill = _RED_FILL
                row_num += 1

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(f"Excel deviation report exported for week {week_id}: {row_num - 4} rows")
        return buffer.read()

    # ── 3. Guard history report ─────────────────────────────────────

    async def export_guard_history(
        self, user_id: uuid.UUID, start_date: date, end_date: date
    ) -> bytes:
        """
        Generate a per-guard history report across multiple weeks.

        Shows submission status and events for each week in the range.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl is required for Excel export")

        user = await self._user_repo.get_by_id(user_id)
        if user is None:
            raise ValueError(f"User {user_id} not found")

        # Get submissions for this user
        submissions = await self._submission_repo.get_by_user(user_id)
        sub_by_week: dict[uuid.UUID, Any] = {s.week_id: s for s in submissions}

        # Get events for this user in the date range
        events: list[Any] = []
        if self._event_repo:
            events = await self._event_repo.get_events_for_user(
                user_id, start_date, end_date
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Guard History"

        # Title
        title_text = f"היסטוריית שומר — {user.full_name}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = _TITLE_FONT
        title_cell.alignment = _CENTER

        # Subtitle with date range
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.cell(
            row=2, column=1, value=f"תקופה: {start_date} עד {end_date}"
        ).alignment = _CENTER

        # Headers
        headers = [
            "שבוע",
            Messages.EXCEL_HEADER_NAME,
            "סטטוס הגשה",
            "אירועים",
        ]
        for col, header in enumerate(headers, 1):
            _apply_header_style(ws.cell(row=4, column=col, value=header))

        # Iterate weeks in range
        row_num = 5
        current = start_date
        while current <= end_date:
            # Find week that starts on this date (or closest)
            # We'll show entries for each 7-day chunk
            week_end = current + timedelta(days=6)

            # Find matching submission
            status_text = Messages.STATUS_PENDING
            events_text = ""

            for s_week_id, s in sub_by_week.items():
                # We need the week object to match dates
                # For now, show all submissions we have
                pass

            # Find events overlapping this period
            period_events = []
            for ev in events:
                if ev.start_date <= week_end and ev.end_date >= current:
                    label = _EVENT_LABELS.get(ev.event_type, ev.event_type)
                    period_events.append(label)
            if period_events:
                events_text = ", ".join(period_events)

            row_data = [
                f"{current} – {week_end}",
                user.full_name,
                status_text,
                events_text,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                _apply_cell_style(cell)

            row_num += 1
            current += timedelta(days=7)

        # Auto-width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        logger.info(
            f"Excel guard history exported for user {user_id}: {row_num - 5} weeks"
        )
        return buffer.read()
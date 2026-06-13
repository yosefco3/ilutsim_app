"""
Tests for Excel export service and controller.
"""

import io
import json
import uuid
from datetime import date, timedelta
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.constants import ShiftType, SubmissionStatus, WeekStatus
from app.messages import Messages
from app.services.excel_export_service import ExcelExportService


# ── Helpers ──────────────────────────────────────────────────────────


def _make_user(user_id=None, full_name="Test Guard", phone="0501234567"):
    user = MagicMock()
    user.id = user_id or uuid.uuid4()
    user.full_name = full_name
    user.phone_number = phone
    user.is_active = True
    return user


def _make_week(week_id=None, week_start=None):
    # Mirror the real ScheduleWeek attributes (start_date/end_date) so the
    # mock catches attribute-name drift. week_start/week_end kept as aliases
    # for the legacy weekly-grid export which still reads them.
    start = week_start or date(2025, 1, 5)  # Sunday
    end = start + timedelta(days=6)
    week = MagicMock()
    week.id = week_id or uuid.uuid4()
    week.start_date = start
    week.end_date = end
    week.week_start = start
    week.week_end = end
    week.status = WeekStatus.OPEN
    return week


def _make_submission(user_id, week_id, status=SubmissionStatus.SUBMITTED):
    sub = MagicMock()
    sub.id = uuid.uuid4()
    sub.user_id = user_id
    sub.week_id = week_id
    sub.status = status
    return sub


def _create_service(
    week=None,
    users=None,
    submissions=None,
):
    """Build an ExcelExportService with mocked repos."""
    sub_repo = AsyncMock()
    user_repo = AsyncMock()
    week_repo = AsyncMock()

    week_repo.get_by_id.return_value = week
    user_repo.get_active_users.return_value = users or []
    sub_repo.get_by_week.return_value = submissions or []
    sub_repo.get_submissions_for_week.return_value = submissions or []
    sub_repo.get_by_user.return_value = submissions or []
    user_repo.get_by_id.return_value = users[0] if users else None

    return ExcelExportService(
        submission_repo=sub_repo,
        user_repo=user_repo,
        week_repo=week_repo,
    )


# ── Weekly schedule export tests ────────────────────────────────────


@pytest.mark.asyncio
async def test_export_weekly_schedule_basic():
    """Basic weekly schedule export with one user and submission."""
    user = _make_user()
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    assert isinstance(data, bytes)
    assert len(data) > 0

    # Verify it's a valid Excel (ZIP magic bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_weekly_no_submission_auto_absence():
    """Users without submissions get auto-absence markers."""
    user = _make_user()
    week = _make_week()

    svc = _create_service(week=week, users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_weekly_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_weekly_schedule(uuid.uuid4())


@pytest.mark.asyncio
async def test_export_weekly_no_openpyxl():
    """Raises RuntimeError when openpyxl is not installed."""
    svc = _create_service(week=_make_week())

    with patch("app.services.excel_export_service.HAS_OPENPYXL", False):
        with pytest.raises(RuntimeError, match="openpyxl"):
            await svc.export_weekly_schedule(svc._week_repo.get_by_id.return_value.id)


# ── Constraints report tests ────────────────────────────────────────


def _make_shift_window(shift_type, start, end):
    sw = MagicMock()
    sw.shift_type = shift_type
    sw.start_time = start
    sw.end_time = end
    return sw


def _make_daily_status(day, is_available, shift_windows=None):
    ds = MagicMock()
    ds.date = day
    ds.is_available = is_available
    ds.shift_windows = shift_windows or []
    return ds


def _make_constraint_submission(user_id, week_id, daily_statuses=None, notes=None):
    sub = _make_submission(user_id, week_id)
    sub.daily_statuses = daily_statuses or []
    sub.general_notes = notes
    return sub


@pytest.mark.asyncio
async def test_export_constraints_basic():
    """Constraints report with one submitting guard."""
    from datetime import time

    user = _make_user(full_name="בני לוי", phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0))],
    )
    sub = _make_constraint_submission(
        user.id, week.id, [ds], notes="זמין רק בבקרים"
    )

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_constraints_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_constraints_report(uuid.uuid4())


@pytest.mark.asyncio
async def test_export_constraints_no_openpyxl():
    """Raises RuntimeError when openpyxl is not installed."""
    svc = _create_service(week=_make_week())

    with patch("app.services.excel_export_service.HAS_OPENPYXL", False):
        with pytest.raises(RuntimeError, match="openpyxl"):
            await svc.export_constraints_report(
                svc._week_repo.get_by_id.return_value.id
            )


@pytest.mark.asyncio
async def test_constraints_excel_has_shift_times_and_notes():
    """Verify the constraints Excel shows shift windows and notes."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    from datetime import time

    user = _make_user(full_name="בני לוי", phone="0502222222")
    week = _make_week(week_start=date(2025, 1, 5))
    ds_sun = _make_daily_status(
        date(2025, 1, 5),
        True,
        [
            _make_shift_window(ShiftType.NIGHT, time(22, 0), time(6, 0)),
            _make_shift_window(ShiftType.MORNING, time(6, 0), time(14, 0)),
        ],
    )
    ds_mon = _make_daily_status(date(2025, 1, 6), False)
    sub = _make_constraint_submission(
        user.id, week.id, [ds_sun, ds_mon], notes="הערה כללית"
    )

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # RTL sheet for Hebrew
    assert ws.sheet_view.rightToLeft is True

    # Header row (row 4): name, phone, period, then days, then notes
    assert ws.cell(row=4, column=1).value == Messages.EXCEL_HEADER_NAME
    assert ws.cell(row=4, column=2).value == Messages.EXCEL_HEADER_PHONE
    assert ws.cell(row=4, column=3).value == "משמרת"
    assert ws.cell(row=4, column=11).value == "הערות"

    # Each guard spans three rows (בוקר / צהריים / ערב), starting at row 5.
    assert ws.cell(row=5, column=1).value == "בני לוי"
    assert ws.cell(row=5, column=3).value == "בוקר"
    assert ws.cell(row=6, column=3).value == "ערב"
    assert ws.cell(row=7, column=3).value == "לילה"

    # Sunday (col 4): morning window in the בוקר row, night window in the לילה row.
    assert "06:00" in ws.cell(row=5, column=4).value
    assert "22:00" in ws.cell(row=7, column=4).value
    # Afternoon (ערב) row has no window that day → empty.
    assert not ws.cell(row=6, column=4).value
    # Monday (col 5): not available → merged "לא זמין" on the top row.
    assert ws.cell(row=5, column=5).value == "לא זמין"
    # Notes merged across the three rows, anchored on the top row.
    assert ws.cell(row=5, column=11).value == "הערה כללית"


@pytest.mark.asyncio
async def test_constraints_excel_has_thick_separator_between_guards():
    """Each guard's three-row block ends with a heavy bottom border so
    adjacent guards are easy to tell apart."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    from datetime import time

    week = _make_week(week_start=date(2025, 1, 5))
    user_a = _make_user(full_name="אבי כהן", phone="0501111111")
    user_b = _make_user(full_name="בני לוי", phone="0502222222")
    ds_a = _make_daily_status(
        date(2025, 1, 5),
        True,
        [_make_shift_window(ShiftType.MORNING, time(7, 0), time(15, 0))],
    )
    ds_b = _make_daily_status(date(2025, 1, 5), False)
    sub_a = _make_constraint_submission(user_a.id, week.id, [ds_a])
    sub_b = _make_constraint_submission(user_b.id, week.id, [ds_b])

    svc = _create_service(
        week=week, users=[user_a, user_b], submissions=[sub_a, sub_b]
    )

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_constraints_report(week.id)

    ws = openpyxl.load_workbook(io.BytesIO(data)).active

    # Guards are sorted by name: אבי (rows 5-7), בני (rows 8-10).
    # openpyxl derives a merged range's bottom edge from its top-left (anchor)
    # cell, so the merged columns carry the thick side on the block's top row;
    # the non-merged period cells carry it on the block's last row.
    def _bottom(row, col):
        return ws.cell(row=row, column=col).border.bottom.style

    # Non-merged period column (3): thick on each block's last (לילה) row.
    assert _bottom(7, 3) == "thick"
    assert _bottom(10, 3) == "thick"
    # Merged name column (1): thick carried on the anchor row of each block.
    assert _bottom(5, 1) == "thick"
    assert _bottom(8, 1) == "thick"
    # Interior period rows are not separators.
    assert _bottom(5, 3) == "thin"
    assert _bottom(6, 3) == "thin"


# ── Deviation report tests ──────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_deviation_report_basic():
    """Deviation report with one guard below threshold."""
    user = _make_user()
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_deviation_missing_submission():
    """Guards with no submission are shown in deviation report."""
    user = _make_user()
    week = _make_week()

    svc = _create_service(week=week, users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    assert isinstance(data, bytes)


@pytest.mark.asyncio
async def test_export_deviation_week_not_found():
    """Raises ValueError for unknown week ID."""
    svc = _create_service()

    with pytest.raises(ValueError, match="not found"):
        await svc.export_deviation_report(uuid.uuid4())


# ── Guard history report tests ──────────────────────────────────────


@pytest.mark.asyncio
async def test_export_guard_history_basic():
    """Guard history report across a date range."""
    user = _make_user()
    start = date(2025, 1, 5)
    end = date(2025, 1, 19)  # 2 weeks

    svc = _create_service(users=[user], submissions=[])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_guard_history(user.id, start, end)

    assert isinstance(data, bytes)
    assert data[:2] == b"PK"


@pytest.mark.asyncio
async def test_export_guard_history_user_not_found():
    """Raises ValueError for unknown user ID."""
    svc = _create_service(users=[])

    with pytest.raises(ValueError, match="not found"):
        await svc.export_guard_history(uuid.uuid4(), date(2025, 1, 5), date(2025, 1, 19))


# ── Integration: verify Excel content via openpyxl ───────────────────


@pytest.mark.asyncio
async def test_weekly_excel_has_correct_structure():
    """Verify the generated Excel has correct title and header rows."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    user = _make_user(full_name="יוסי כהן")
    week = _make_week(week_start=date(2025, 1, 5))
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_weekly_schedule(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # Title row
    assert ws.cell(row=1, column=1).value is not None

    # Header row (row 3)
    assert ws.cell(row=3, column=1).value == Messages.EXCEL_HEADER_NAME
    assert ws.cell(row=3, column=2).value == "ראשון"

    # Data row (row 4) - user name
    assert ws.cell(row=4, column=1).value == "יוסי כהן"


@pytest.mark.asyncio
async def test_deviation_excel_has_correct_data():
    """Verify deviation Excel shows guards below threshold."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not installed")

    user = _make_user(full_name="דני לוי", phone="0509876543")
    week = _make_week()
    sub = _make_submission(user.id, week.id)

    svc = _create_service(week=week, users=[user], submissions=[sub])

    with patch("app.services.excel_export_service.HAS_OPENPYXL", True):
        data = await svc.export_deviation_report(week.id)

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active

    # Header row
    assert ws.cell(row=3, column=1).value == Messages.EXCEL_HEADER_NAME

    # Data row - should have user with submission status
    assert ws.cell(row=4, column=1).value == "דני לוי"
